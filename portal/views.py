from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import views as auth_views
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.utils import timezone
from django.db import IntegrityError
from django.http import HttpResponse, HttpResponseNotFound, HttpResponseServerError
from django.template.loader import get_template
from xhtml2pdf import pisa
import threading
from .utils import *
from .models import *
from .forms import AuthForm
import openpyxl
from openpyxl.styles import Font, PatternFill,Border, Side
from .models import Student, Record, Document


class LoginView(auth_views.LoginView):
    form_class = AuthForm


def home(request):
    return render(request, "dashboard.html")

@login_required
def add(request):
    if request.method == "POST":
        
        # Getting the data from post request
        post_data = request.POST.dict()
        quota = post_data['quota'] == 'govt'
        
        # Creating an entry for the student admission num on first submit
        try:
            student = Student(admission_no=post_data['receipt'],
                            version_count=0,
                            lock=False)
            student.save()

        # If a student of given id already exists then showing an error message
        except IntegrityError:
            return render(request, "dup_index.html", {"admission_no": post_data['receipt']})
    
        version = Version(student=student,
                          version_count=0,
                          stud_ver=0,
                          docs_ver=0,
                          curr_user=User.objects.get(username = request.user))
        
        version.save()

        # Saving the student info
        student_info = StudentInfo(
                              name=post_data['name_stu'],
                              student = student,
                              parent_name=post_data['name_prnt'],
                              department=post_data['dept'],
                              student_number=post_data['contact1'],
                              parent_number=post_data['contact2'],
                              quota=quota,
                              email=post_data['email'],
                              ver=0)
        student_info.save()

        # Getting all filenames from the form
        file_names = [[*name.split(':')] for name in post_data.keys() if ":" in name]
        clean_names = {}
        for name in file_names:
            if name[0] in clean_names:
                clean_names[name[0]].append(name[1])
            else:
                clean_names[name[0]] = [name[1]]
        
        # Saving the file data
        for file in clean_names:
            doc = Document.objects.get(name = file)

            # Getting the info from post request
            original = post_data[file+":original"] == 'on'
            photo_copy = post_data[file+":copy"] == 'on'
            count = int(post_data[file+":count"])
            
            Record(student=student, document=doc,
                    original=original,
                    photocopy=photo_copy,
                    count=count,
                    date=timezone.localdate(),
                    ver=0).save()
            
        # Mailing the student
        context = get_student_info(request, student.admission_no)
        context['cur_ver'] += 1
        
        threading.Thread(target=mail_student, args=("stud_mail.html", context, student_info.email)).start()
            
        return redirect('view', admission_no=student.admission_no)
    else:

        file_names = [document.name for document in Document.objects.all()]
        return render(request, "index.html", {"file_names": file_names})
        

@login_required
def edit(request, admission_no):

    student = get_object_or_404(Student, admission_no=admission_no)
    version = get_object_or_404(Version, version_count=student.version_count, student=student)
    info = get_object_or_404(StudentInfo, student=student, ver=version.stud_ver)
    records = Record.objects.filter(student=student, ver=version.docs_ver)
    version_values = [i + 1 for i in range(0, student.version_count)]

    # Checking if the document is locked
    if student.lock:
        return render(request, "lock.html")

    if request.method == "POST":

        # Checking if the document must be locked
        if request.POST.get("lock", False):
            student.lock = True
            student.save()

        # Getting the data from post request
        post_data = request.POST.dict()
        quota = post_data['quota'] == 'govt'
        
        # Checking for new changes in student details
        stud_changed = False in [info.name == post_data['name_stu'],
        info.parent_name == post_data['name_prnt'],
        info.department == post_data['dept'],
        info.student_number == post_data['contact1'],
        info.parent_number == post_data['contact2'],
        info.quota == quota,
        info.email == post_data["email"]]

        # Getting all filenames from the form
        file_names = [[*name.split(':')] for name in post_data.keys() if ":" in name]
        clean_names = {}
        for name in file_names:
            if name[0] in clean_names:
                clean_names[name[0]].append(name[1])
            else:
                clean_names[name[0]] = [name[1]]
        
        # Checking for changes in documents
        doc_changed = False
        for doc in clean_names:
            original = post_data[f"{doc}:original"] == 'on'
            copy = post_data[f"{doc}:copy"] == 'on'
            count = int(post_data[f"{doc}:count"])
            rec = records.get(document__name = doc)
            if False in [rec.original==original,
                    rec.photocopy==copy,
                    rec.count==count]:
                doc_changed = True
                break
        
        # Checking for modification
        if doc_changed or stud_changed:

             # incrementing the version count
            student.version_count += 1
            student.save()

            # Creating a new version object to keep track of change
            new_version = Version(student=student,
                            version_count=student.version_count,
                            stud_ver=version.stud_ver,
                            docs_ver=version.docs_ver,
                            curr_user=User.objects.get(username= request.user))
            

            if stud_changed:
                new_version.stud_ver += 1
                new_stud = StudentInfo(
                                name=post_data['name_stu'],
                                student = student,
                                parent_name=post_data['name_prnt'],
                                department=post_data['dept'],
                                student_number=post_data['contact1'],
                                parent_number=post_data['contact2'],
                                quota=quota,
                                email=post_data["email"],
                                ver=new_version.stud_ver)
                new_stud.save()

            if doc_changed:

                new_version.docs_ver += 1
                
                # Saving the file data
                for file in clean_names:

                    doc = Document.objects.get(name = file)
                    rec = records.get(document__name = doc)
                    
                    # Getting the info from post request
                    original = post_data[file+":original"] == 'on'
                    photo_copy = post_data[file+":copy"] == 'on'
                    count = int(post_data[file+":count"])
                    
                    # Checking if this specific file was modified
                    date = rec.date
                    if False in [rec.original==original,
                    rec.photocopy==copy,
                    rec.count==count]:
                        date = timezone.localtime()
                    Record(student=student, document=doc,
                            original=original,
                            photocopy=photo_copy,
                            count=count,
                            date=date,
                            ver=new_version.docs_ver).save()
            
            new_version.save()
    
            # Mailing the student
            context = get_student_info(request, student.admission_no)
            context['cur_ver'] += 1

            threading.Thread(target=mail_student, args=("stud_mail.html", context, post_data["email"])).start()
                
        return redirect('view', admission_no=student.admission_no)

    departments = ["CSE", "ECE", "CSBS", "AI&DS", "MECH", "IT", "MBA", "CYS", "AI&ML"]
    return render(request, "edit.html", {"student": info, "records": records, "admission_no": admission_no, "versions": version_values, "cur_ver": version.version_count, 'depts': departments})

@login_required
def view(request, admission_no):

    context = get_student_info(request, admission_no)
    return render(request, "next-page.html", context)

@login_required
def pdf_download(request, admission_no):

    template_path = 'print.html'
    
    context = get_student_info(request, admission_no)
    
    records, extra, serial, extra_index = split_records(context["records"])
    records = zip(records, serial)
    context["records1"] = records
    
    records, extra, serial, extra_index = split_records(context["records"])
    records = zip(records, serial)
    
    context["records2"] = records
    context["extra"] = extra
    context["extra_index"] = extra_index
    
    # Create a Django response object, and specify content_type as pdf
    response = HttpResponse(content_type='application/pdf')
    
    # find the template and render it.
    template = get_template(template_path)
    html = template.render(context)

    # create a pdf
    pisa_status = pisa.CreatePDF(
       html, dest=response)
    # if error then show some funny view
    if pisa_status.err:
       return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response

@login_required
def stud(request, admission_no):
    
    context = get_student_info(request, admission_no)
    context['cur_ver'] += 1
    mail_student("stud_mail.html", context, "marudhu2021@gmail.com")
    return render(request, "stud_mail.html", context)

class LoginView(auth_views.LoginView):
    form_class = AuthForm

@login_required
def export_certificate_summary(request):
    # Create an in-memory workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Student Certificate Summary'

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Add headers for student details and dynamically add columns for certificates
    documents = Document.objects.all()
    headers = [
        'Student Name', 
        'Student Admission No', 
        'Department', 
        'Student Contact Number', 
        'Parent Contact Number', 
        'Quota',  
        'Version Count'
    ]

    # Add dynamic columns (Original/Copy) for each document
    for doc in documents:
        headers.append(f'{doc.name} (Original)')
        headers.append(f'{doc.name} (Photocopy)')

    # Append headers
    sheet.append(headers)

    # Style the headers (bold and bordered)
    for cell in sheet["1:1"]:
        cell.font = Font(bold=True)
        cell.border = thin_border

    # Fetch all students
    students = Student.objects.all()

    # Iterate over each student and add their information along with document submissions
    for student in students:
        # Fetch student info
        student_info = StudentInfo.objects.filter(student=student).first()

        # Prepare student details
        student_name = student_info.name if student_info else 'Unknown'
        department = student_info.department if student_info else 'Unknown'
        student_contact = student_info.student_number if student_info else 'Unknown'
        parent_contact = student_info.parent_number if student_info else 'Unknown'
        quota = 'Government' if student_info and student_info.quota else 'Management'
        version_count = student.version_count + 1  # Start version count at 1 instead of 0

        # Add student details to the row
        row = [
            student_name, 
            student.admission_no, 
            department, 
            student_contact, 
            parent_contact, 
            quota, 
            version_count
        ]

        # For each student, check if they have submitted each certificate (original and photocopy)
        for document in documents:
            record = Record.objects.filter(student=student, document=document).first()

            # Original Document
            if record and record.original:
                original_status = f'Submitted'
                original_color_fill = PatternFill(start_color="5cef65", end_color="5cef65", fill_type="solid")  # Green
            else:
                original_status = 'Not Submitted'
                original_color_fill = PatternFill(start_color="ee3f19", end_color="ee3f19", fill_type="solid")  # Red

            # Photocopy Document
            if record and record.photocopy:
                photocopy_status = f'Submitted (Qty: {record.count})'
                photocopy_color_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")  # Blue
            else:
                photocopy_status = 'Not Submitted'
                photocopy_color_fill = PatternFill(start_color="ee3f19", end_color="ee3f19", fill_type="solid")  # Red

            # Append both original and photocopy statuses to the row
            row.append(original_status)
            row.append(photocopy_status)

        # Append the row to the sheet
        sheet.append(row)

        # Apply color immediately after adding the row
        # Starting from the 8th column (which dynamically shifts as more documents are added)
        col_offset = len(headers) - (2 * len(documents))  # To handle column offset dynamically
        col_offset = 8  # Assume the document-related columns start at column 8, adjust as necessary
        for idx, document in enumerate(documents):
            record = Record.objects.filter(student=student, document=document).first()  # Ensure the record is specific to this document

    # Original Document cell
            original_cell = sheet.cell(row=sheet.max_row, column=col_offset + idx * 2)
            if record and record.original:
                original_cell.fill = PatternFill(start_color="5cef65", end_color="5cef65", fill_type="solid")  # Green
            else:
                original_cell.fill = PatternFill(start_color="ee3f19", end_color="ee3f19", fill_type="solid")  # Red
            original_cell.border = thin_border

    # Photocopy Document cell
            photocopy_cell = sheet.cell(row=sheet.max_row, column=col_offset + idx * 2 + 1)
            if record and record.photocopy:
                photocopy_cell.fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")  # Blue
            else:
                photocopy_cell.fill = PatternFill(start_color="ee3f19", end_color="ee3f19", fill_type="solid")  # Red
            photocopy_cell.border = thin_border
    # Apply borders to all cells
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border

    # Create a response object for the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=certificate_summary.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response



######################## error pages ########################

def error_404(request, exception):
    return HttpResponseNotFound(render(request, 'error_404.html'))

def error_500(request):
    return HttpResponseServerError(render(request, 'error_500.html'))